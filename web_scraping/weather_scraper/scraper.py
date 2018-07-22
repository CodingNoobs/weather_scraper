#! python3
import requests, datetime, pprint, os.path, time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.compat import range

# Try to run program.
url = 'https://forecast.weather.gov/MapClick.php?lat=39.59&lon=-105.01'

# Infinite loop.
while True:
    try:
        # Request given URL.
        r = requests.get(url, timeout=5)
        
        # Create BeautifulSoup object with the .text of our request object.
        soup = BeautifulSoup(r.text, 'html.parser')

    # Handle exceptions.
    except requests.ConnectionError as e:
        print("Connection Error. Make sure you are connected to Internet. Technical Details given below.\n")
        print(str(e))
    except requests.Timeout as e:
        print("Timeout Error")
        print(str(e))
    except requests.RequestException as e:
        print("General Error")
        print(str(e))
    except KeyboardInterrupt:
        print("Someone closed the program")
    finally:

        title_list = ['date', 'time', 'location', 'latitude', 'longitude', 'elevation', 'predicted_high_temp', 'current_temp']
        # Create dictionary to hold weather data.
        weather = {}

        # Set variables in dict.
        weather['date'] = datetime.date.today()
        weather['time'] = datetime.datetime.now().time()

        # Safely get variables that have been scraped from HTML.
        try:
            weather['location_name'] = soup.find('h2', class_='panel-title').string
        except AttributeError:
            weather['location_name'] = 'no data'

        try:
            weather['latitude'] = soup.find('span', class_='smallTxt').contents[1].string
        except AttributeError:
            weather['latitude'] = 'no data'

        try:
            weather['longitude'] = soup.find('span', class_='smallTxt').contents[3].string
        except AttributeError:
            weather['longitude'] = 'no data'

        try:
            weather['elevation'] = soup.find('span', class_='smallTxt').contents[5].string
        except AttributeError:
            weather['elevation'] = 'no data'

        try:
            weather['predicted_high_temp'] = soup.find('p', class_='temp-high').string
        except AttributeError:
            weather['predicted_high_temp'] = 'no data'

        try:
            weather['current_temp'] = soup.find('p', class_='myforecast-current-lrg').string
        except AttributeError:
            weather['current_temp'] = 'no data'

        # Print out the key value pairs
        for key, value in weather.items():
            print('{}: {}'.format(key, value))

        # The name of the file.
        dest_filename = 'denver_weather.xlsx'

        # Check if the dest_file already exists.
        if not os.path.exists(dest_filename):
            print('File doesn\'t exist!')

            # Create a new work.
            wb = Workbook()

            # Create a new sheet.
            ws1 = wb.active

            # Name the new sheet.
            ws1.title = "Denver_Weather"
            
            # Index 0 won't work in Excel
            for row in range(1, 2):
                for col in range(len(title_list)):
                    # Add 1 to column index to avoid index 0 problem in Excel.
                    ws1.cell(column=col + 1, row=row, value="{}".format(title_list[col]))

        else:
            print('File exists!')

            # Load the already created workbook.
            wb = load_workbook(filename = dest_filename)
            ws1 = wb.active

            # Create a dict with keys eqaul to the spreadsheet columns.
            excel_dict = {'A': str(weather['date']),
                        'B': weather['time'],
                        'C': str(weather['location_name']),
                        'D': str(weather['latitude']),
                        'E': str(weather['longitude']),
                        'F': str(weather['elevation']),
                        'G': str(weather['predicted_high_temp']),
                        'H': str(weather['current_temp'])}
            
            ws1.append(excel_dict)

        # Save the file.
        wb.save(filename = dest_filename)

        # Wait x amount of time before next iteration.
        time.sleep(3600)