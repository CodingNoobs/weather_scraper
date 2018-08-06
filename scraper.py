#! python3

# This program scrapes data from a weather website and saves into to an Excel doc.

# Import external Python packages used by this program.
import requests, datetime, pprint, os.path, time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.compat import range

# The URL of the website we want to scrape.
url = 'https://forecast.weather.gov/MapClick.php?lat=39.59&lon=-105.01'

# The name of the file to save the data to.
dest_filename = 'denver_weather.xlsx'

# Infinite loop. Runs the program until stopped. Use CTRL+C to stop.
while True:

    # Try to get the requested URL. 
    try:
        # Request given URL. Timeout is how long a request will hang before giving up.
        r = requests.get(url, timeout=5)

    # Handle exceptions.
    except requests.ConnectionError as e:
        print("Connection Error. Make sure you are connected to Internet. Technical Details given below.\n")
        print(str(e))
    except requests.Timeout as e:
        print("Timeout Error")
        print(str(e))
    except requests.RequestException as e:
        print("General Error")
        print(str(e))
    except KeyboardInterrupt:
        print("Someone closed the program")

    # Do this if the program throws no exceptions.
    else:

        # Create BeautifulSoup object with the .text of our request object.
        soup = BeautifulSoup(r.text, 'html.parser')

        # Create dictionary to hold weather data.
        weather = {}

        # Set non-scraping variables in dict.
        weather['date'] = datetime.date.today()
        weather['time'] = datetime.datetime.now().time()

        # Safely get variables that have been scraped from HTML.
        try:
            weather['location_name'] = soup.find('h2', class_='panel-title').string
        except AttributeError:
            weather['location_name'] = 'no data'

        try:
            weather['latitude'] = soup.find('span', class_='smallTxt').contents[1].string
        except AttributeError:
            weather['latitude'] = 'no data'

        try:
            weather['longitude'] = soup.find('span', class_='smallTxt').contents[3].string
        except AttributeError:
            weather['longitude'] = 'no data'

        try:
            weather['elevation'] = soup.find('span', class_='smallTxt').contents[5].string
        except AttributeError:
            weather['elevation'] = 'no data'

        try:
            weather['predicted_high_temp'] = soup.find('p', class_='temp-high').string
        except AttributeError:
            weather['predicted_high_temp'] = 'no data'

        try:
            weather['current_temp'] = soup.find('p', class_='myforecast-current-lrg').string
        except AttributeError:
            weather['current_temp'] = 'no data'

        # Print out the key value pairs of the weather dict to terminal.
        # for key, value in weather.items():
        #     print('{}: {}'.format(key, value))

        # Check if the dest_file already exists.
        if not os.path.exists(dest_filename):
            print('File doesn\'t exist!')

            # Create a new Excel workbook.
            wb = Workbook()

            # Create a new sheet.
            ws1 = wb.active

            # Name the new sheet.
            ws1.title = "Denver_Weather"

            # A list that holds the names of the columns for the Excel document.
            title_list = ['date', 'time', 'location', 'latitude', 'longitude', 'elevation', 'predicted_high_temp', 'current_temp']
            
            # Add the column name list to the sheet.
            # Add to row 1.
            for row in range(1, 2):
                # Add the columns to worksheet for the length of the title_list.
                for col in range(len(title_list)):
                    # Add 1 to column index to avoid index 0 problem in Excel.
                    ws1.cell(column=col + 1, row=row, value="{}".format(title_list[col]))

            # Create a dict with keys equal to the spreadsheet columns.
            excel_dict = {'A': str(weather['date']),
                          'B': weather['time'],
                          'C': str(weather['location_name']),
                          'D': str(weather['latitude']),
                          'E': str(weather['longitude']),
                          'F': str(weather['elevation']),
                          'G': str(weather['predicted_high_temp']),
                          'H': str(weather['current_temp'])}

            # Append the next row to the Excel doc.
            ws1.append(excel_dict)

        # Otherwise, just add data to the file since it exists.
        else:
            print('File exists!')

            # Load the already created workbook.
            wb = load_workbook(filename = dest_filename)
            ws1 = wb.active

            # Just add values that change to next row.
            excel_dict = {'A': str(weather['date']),
                          'B': weather['time'],
                          'G': str(weather['predicted_high_temp']),
                          'H': str(weather['current_temp'])}
            
            # Append the next row to the Excel doc.
            ws1.append(excel_dict)

        # Save the file.
        wb.save(filename = dest_filename)
        print('File saved!')

        # Wait time. Don't scrape more than 1 request per second. Number is in seconds.
        wait_time = 3600

        # Wait x amount of time before next iteration.
        time.sleep(wait_time)